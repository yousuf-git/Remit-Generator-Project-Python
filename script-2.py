import time
import os
import shutil # for copying file from one location to other
import datetime as dt
import openpyxl as xl
from openpyxl.styles import Font, Border, Side, Alignment;
from tqdm import tqdm # for loading-bar
import threading

def loading_screen(duration=20, iterations=230):
    sleep_time = duration / iterations
    for _ in tqdm(range(iterations), desc="", ncols=75):
        time.sleep(sleep_time)

thin_boarder = Border (
    left = Side(style = "thin"),
    right = Side(style = "thin"),
    top = Side(style = "thin"),
    bottom = Side(style = "thin")
)
font = Font(size=9)

# Get the directory of the current script
script_dir = os.path.dirname(os.path.abspath(__file__))

# List all Excel files in the directory
for filename in os.listdir(script_dir):
    if filename.endswith('.xlsx'):
        if filename.__contains__("helper"):
            helperPath = os.path.join(script_dir, filename)
        elif filename.__contains__("Template"):
            templatePath = os.path.join(script_dir, filename)

'''Load Helper Workbook'''

print("Helper Loading.....")
loading_thread = threading.Thread(target=loading_screen)
loading_thread.start()
helperWb = xl.load_workbook(filename=helperPath, data_only=True)
loading_thread.join()
print("\nHelper Loaded")
helperSheet = helperWb.active

"""Load Template Workbook"""
print("Template Loading.....")
templateWb = xl.load_workbook(filename=templatePath)
print("Template Loaded")
templatePaymentsSheet = templateWb["""Payments & NSF's"""]

for i in range(2, templatePaymentsSheet.max_row+1):
    cell = helperSheet[f'K{i}']
    if cell.value == "SIF":
        templatePaymentsSheet[f'K{i}'].value = "SIF"

print("SIF values moved to Template.")

for i in range(2, templatePaymentsSheet.max_row - 1):
    templatePaymentsSheet[f'K{i}'].border = thin_boarder
    templatePaymentsSheet[f'K{i}'].font = font
    templatePaymentsSheet[f'K{i}'].alignment = Alignment(horizontal = 'right')

print("Template Saving...")
templateWb.save(templatePath)
print("Template Saved")

'''Copy Final Template from here to date folder'''

curr_dir = os.path.dirname(os.path.abspath(__file__))
curr_year = dt.datetime.now().year
curr_date = dt.datetime.now().strftime('%m-%d-%Y')

year_path = os.path.join(curr_dir, str(curr_year))
date_path = os.path.join(year_path, curr_date)

if not os.path.exists(year_path):
    os.makedirs(year_path)
if not os.path.exists(date_path):
    os.makedirs(date_path)

srcFile = ""
for filename in os.listdir(curr_dir):
    if filename.__contains__("Template"):
        srcFile = filename
        
srcPath = os.path.join(curr_dir, srcFile)
destPath = os.path.join(curr_dir, year_path, date_path, f"Weekly Remit {curr_date.replace("-", "")}.xlsx")
shutil.copy(srcPath, destPath)

destPath = os.path.join(curr_dir, year_path, date_path, f"Weekly Remit {curr_date.replace("-", "")}.xls")
os.rename(srcPath, destPath)