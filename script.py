import os
import openpyxl as xl # for excel files modification
from openpyxl.styles import Border, Side, Font 
import pandas as pd # for data analysis like grouoing
import time
from tqdm import tqdm # for loading bar generation
import threading # for creating thread of loading-bar function while loading a file

# loading bar for 20 seconds
def loading_screen(duration=20, iterations=230):
    sleep_time = duration / iterations
    for _ in tqdm(range(iterations), desc="", ncols=75):
        time.sleep(sleep_time)

# Define the border style
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Define the font style (bold text)
bold_font = Font(size=10, bold=True)

# To copy cells from one sheet to other in the given range
def copyCells(src_sheet, src_range, dest_sheet, dest_start_cell):
    """
    This method is to copy range of cells from one sheet to other
    """
    # Parse the source range into cells (given will be in the form of string "A1:C3")
    src_range_cells = src_sheet[src_range]
    # print("Source Range:", src_range_cells)

    # Calculate starting row and column for destination
    dest_start_row = dest_start_cell[1]
    dest_start_col = dest_start_cell[0]

    # # Copy the range of cells
    print("Copying Started.....")
    for i, row in enumerate(src_range_cells): # i => 0 to len(src_range_cells) - 1
        print(".", end="")
        for j, cell in enumerate(row): # j => 0 to len(row) - 1
            # dest_sheet.cell(row = dest_start_row + i, column = dest_start_col + j).value = cell.value
            # dest_sheet.cell(row = dest_start_row + i, column = dest_start_col + j , value = cell.value)
            destCell = dest_sheet.cell(row = dest_start_row + i, column = dest_start_col + j)
            destCell.value = cell.value
            destCell.border = thin_border
            destCell.font = Font(size=9)
    print("Copying done.")

'''Declaring Path of all sheets'''
# remitPath = "Testings/Test 3/Weekly Remit 05282024.xlsx"
# helperPath = "Testings/Test 3/helper.xlsx"
# templatePath = "Testings/Test 3/Template.xlsx"

# Get the directory of the current script
script_dir = os.path.dirname(os.path.abspath(__file__))

# List all Excel files in the directory
for filename in os.listdir(script_dir):
    if filename.endswith('.xlsx'):
        if filename.__contains__("Remit"):
            remitPath = os.path.join(script_dir, filename)
        elif filename.__contains__("helper"):
            helperPath = os.path.join(script_dir, filename)
        elif filename.__contains__("Template"):
            templatePath = os.path.join(script_dir, filename)
            

"""Load Remit Workbook"""
print("Remit Loading.....")
remitWb = xl.load_workbook(filename=remitPath)
print("Remit Loaded")
remitSheet = remitWb.active

"""Load Helper Workbook"""
print("Helper Loading.....")
loading_thread = threading.Thread(target=loading_screen)
loading_thread.start()
helperWb = xl.load_workbook(filename=helperPath)
loading_thread.join()
print("\nHelper Loaded")
helperSheet = helperWb.active

"""Load Template Workbook"""
print("Template Loading.....")
templateWb = xl.load_workbook(filename=templatePath)
print("Template Loaded")
# print(templateWb.sheetnames)
# print(templateWb["""Payments & NSF's"""])
templatePaymentsSheet = templateWb["""Payments & NSF's"""]

"""Copy 1st part from remit to helper"""
# Destination start cell should be specified as a string and converted to a cell coordinate
src_range = 'A2:J' + str(remitSheet.max_row)
dest_start_cell_str = 'A2'
dest_start_cell = xl.utils.cell.coordinate_from_string(dest_start_cell_str) # ('A', 2)
dest_start_cell = xl.utils.cell.column_index_from_string(dest_start_cell[0]), int(dest_start_cell[1]) # (1, 2)
copyCells(remitSheet, src_range, helperSheet, dest_start_cell)
print("Copying Process 1 Done.")

"""To remove white spaces from cells of col M in Remit"""
colM = remitSheet['M']
for cell in colM:
    if cell.value == " ":
        cell.value = ""
print("White Spaces Removed ")

"""Copy 2nd part from remit to helper"""
src_range = 'L2:P' + str(remitSheet.max_row)
dest_start_cell_str = 'L2'
dest_start_cell = xl.utils.cell.coordinate_from_string(dest_start_cell_str) # ('L', 2)
dest_start_cell = xl.utils.cell.column_index_from_string(dest_start_cell[0]), int(dest_start_cell[1]) # (12, 2)

copyCells(remitSheet, src_range, helperSheet, dest_start_cell)
print("Copying Process 2 Done.")

# helperSheet['K2'].value = '''=IF([@[Total Received]]<[@[SIF Amount]],"",IF([@[SIF/CO]]="SIF","SIF",IF([@[SIF/CO]]="","","SPR")))'''

"""Copy Column A to J from helper to template"""

# print("Helper Sheet 1st SIF Value: ", helperTempSheet['K4'].value, helperTempSheet['K5'].value, helperTempSheet['K6'].value, helperTempSheet['K7'].value, helperTempSheet['K8'].value)

src_range = 'A2:J' + str(remitSheet.max_row)
print("Range =", src_range)
dest_start_cell_str = 'A2'
dest_start_cell = xl.utils.cell.coordinate_from_string(dest_start_cell_str) # ('A', 2)
dest_start_cell = xl.utils.cell.column_index_from_string(dest_start_cell[0]), int(dest_start_cell[1]) # (1, 2)
# print("...////")
copyCells(helperSheet, src_range, templatePaymentsSheet, dest_start_cell)
print("Copying Process 3 Done.")

print("Saving Helper.....")
loading_thread = threading.Thread(target=loading_screen)
loading_thread.start()
helperWb.save(filename=helperPath)
print("\nHelper Saved !")

"""To copy PIF values from remit to Template"""

# Traverse column 'M' in the remit sheet
for row in range(2, remitSheet.max_row + 1): # +1 is exclusive in range()
    cell_value = remitSheet[f'K{row}'].value
    if cell_value == "PIF":
        # Store "PIF" in the same cell location in the template payments sheet
        templatePaymentsSheet[f'K{row}'].value = "PIF"

print("PIF cells copied from remit to template.")        

maxRow = templatePaymentsSheet.max_row

def changeFormat(sheet, range_str, format):
    range_cells = sheet[range_str]
    
    for row in range_cells:
        for cell in row:
            cell.number_format = format

accountingFormat = '$* #,##0.00'
changeFormat(templatePaymentsSheet, f"A2:C{templatePaymentsSheet.max_row}", '@') # Text Format
changeFormat(templatePaymentsSheet, f"E2:G{templatePaymentsSheet.max_row}", accountingFormat)
changeFormat(templatePaymentsSheet, f"H2:H{templatePaymentsSheet.max_row}", '0') # number format
changeFormat(templatePaymentsSheet, f"J2:J{templatePaymentsSheet.max_row}", accountingFormat)
changeFormat(templatePaymentsSheet, f"K2:K{templatePaymentsSheet.max_row}", '@') # Text Format

"""Calculating Sum of Required Columns in Template Payment Sheet"""
# colListforSum= ['E', 'F', 'G', 'J']

print("Max Rows in Template:", maxRow) # for debugging

currency_format = '$#,##0.00'

paySumCell = templatePaymentsSheet[f'E{maxRow+2}']
paySumCell.value = f'=SUM(E2:E{maxRow})'
paySumCell.border = thin_border
paySumCell.font = bold_font
paySumCell.number_format = currency_format

feeSumCell = templatePaymentsSheet[f'F{maxRow+2}']
feeSumCell.value = f'=SUM(F2:F{maxRow})'
feeSumCell.border = thin_border
feeSumCell.font = bold_font
feeSumCell.number_format = currency_format

dueSumCell = templatePaymentsSheet[f'G{maxRow+2}']
dueSumCell.value = f'=SUM(G2:G{maxRow})'
dueSumCell.border = thin_border
dueSumCell.font = bold_font
dueSumCell.number_format = currency_format

blcSumCell = templatePaymentsSheet[f'J{maxRow+2}']
blcSumCell.value = f'=SUM(J2:J{maxRow})'
blcSumCell.border = thin_border
blcSumCell.font = bold_font
blcSumCell.number_format = currency_format

print("Sum of Balances is Stored.")

"""Calculate Individual Sum for each account"""

# Assuring that template workbook is already loaded

# Read the Excel file into a pandas DataFrame
# df = pd.read_excel(file_path, sheet_name=sheet_name)
df = pd.read_excel(remitPath, sheet_name = "CavalryAgencyRemit")

# Calculating the sum of balances for each account holder
due_sum = df.groupby('Current Creditor')['Due Client'].sum().reset_index() # also reset the grouping

# Converting the DataFrame to a dictionary
due_dict = due_sum.set_index('Current Creditor')['Due Client'].to_dict()
print("Distributed Sums =", due_dict) # for debug

"""Store Sum Values in Corresponding cells in template summary sheet"""

# Accessing the template summary sheet
templateSummarySheet = templateWb["Portfolio Summary"]

# C7 -> CAVALRY INVESTMENTS, LLC
# C8 -> CAVALRY SPV I, LLC
# C9 -> CAVALRY SPV II, LLC
# C10 -> CAVALRY PORTFOLIO SERVICES, LLC

print("Creditors Accounts Found:", due_dict.keys()) # for debug

for key in due_dict.keys():
    if key == "CAVALRY INVESTMENTS, LLC":
        templateSummarySheet['C7'].value = due_dict.get(key)
    elif key == "CAVALRY SPV I, LLC":
        templateSummarySheet['C8'].value = due_dict.get(key)
    elif key == "CAVALRY SPV II, LLC":
        templateSummarySheet['C9'].value = due_dict.get(key) 
    elif key == "CAVALRY PORTFOLIO SERVICES, LLC":
        templateSummarySheet['C10'].value = due_dict.get(key)

print("Creditors Due Balance is Stored Individually.")
templateSummarySheet['C11'].value = "=SUM(C7:C10)"
# blcSumCell.number_format = currency_format

"""Saving Template workbook"""

print("Saving Template.....")
templateWb.save(filename=templatePath)
print("Template Saved !")

print("Open and Save the Helper Sheet then execute script-2.")
# input()  # This will keep the terminal open until you press Enter